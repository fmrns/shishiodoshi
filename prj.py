#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Released under MIT License
#
# Copyright (c) 2025 Fumiyuki Shimizu
# Copyright (c) 2025 Abacus Technologies, Inc.
#
# Permission is hereby granted, free of charge, to any person
# obtaining a copy of this software and associated documentation files
# (the "Software"), to deal in the Software without restriction,
# including without limitation the rights to use, copy, modify, merge,
# publish, distribute, sublicense, and/or sell copies of the Software,
# and to permit persons to whom the Software is furnished to do so,
# subject to the following conditions:
#
# The above copyright notice and this permission notice shall be
# included in all copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND,
# EXPRESS OR IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF
# MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND
# NONINFRINGEMENT. IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS
# BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN
# ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN
# CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import sys
import os

sys.path.insert(0, os.path.dirname(__file__))

import subprocess
from datetime import datetime as datetime_sucks
from pendulum import (
    DateTime,
    datetime,
    timezone,
    set_local_timezone,
    now,
    parse as penparse,
    SATURDAY,
    SUNDAY,
)
from openpyxl import load_workbook
import plotly.figure_factory as ff


from util.text import wljustify, TERM_NORM, TERM_RED
from timerange import TimeRange, TimeRangeSet
from task import Task, TaskSet
from member import Member, MemberSet

if os.name == "nt":
    subprocess.run("c:/Windows/System32/mode.com con cp select=65001", shell=True)
    # print(f"sys.stdout.encoding: {sys.stdout.encoding}")

tz_default = timezone("Asia/Tokyo")
excel_epoch = datetime(1899, 12, 30, tz=tz_default)
set_local_timezone(tz_default)


def to_datetime(val):
    if val is None:
        return val
    if isinstance(val, DateTime):
        return val
    if isinstance(val, datetime_sucks):
        return datetime(
            val.year,
            val.month,
            val.day,
            val.hour,
            val.minute,
            val.second,
            tz=tz_default,
        )
    if isinstance(val, str):
        try:
            return datetime.from_format(val, "YYYY-MM-DD HH:mm:ss", tz=tz_default)
        except Exception:
            return None
    if isinstance(val, int):
        return excel_epoch.add(days=val)
    return val


# O例以降で同じ日時が縦に並んでいる場所の(i+1行j列)を返す
def find_double_datetime(ws) -> tuple[int, int]:
    rows = list(ws.iter_rows())
    # O列 14(0base)
    for j in range(15 - 1, len(rows[0])):
        for i in range(len(rows) - 2):
            c = to_datetime(rows[i][j].value)
            if isinstance(c, DateTime) and c == to_datetime(rows[i + 1][j].value):
                return i + 2, j + 1  # 1base
    return None, None


def make_breaks(on_off_map: dict[DateTime, bool]):
    the_first = min(on_off_map.keys())
    the_last = max(on_off_map.keys()).add(days=1)

    break_start1 = the_first.add(days=-1).at(17)
    break_end1 = the_first.at(9)
    break_start2 = the_first.at(12)
    break_end2 = the_first.at(13)

    breaks = TimeRangeSet([TimeRange(the_first.add(days=-1).at(0), the_first.at(0))])
    # # 祝日 -> xlsx
    # breaks.add(TimeRange(datetime(2025, 12, 31, 0), datetime(2026, 1, 1, 0)))
    # breaks.add(TimeRange(datetime(2026, 1, 1, 0), datetime(2026, 1, 2, 0)))
    # breaks.add(TimeRange(datetime(2026, 1, 12, 0), datetime(2026, 1, 13, 0)))
    while break_start1 < the_last:
        breaks.add(TimeRange(break_start1, break_end1))
        breaks.add(TimeRange(break_start2, break_end2))
        if break_end1.at(0) >= the_last or not on_off_map[break_end1.at(0)]:
            breaks.add(TimeRange(break_end1.at(0), break_end1.add(days=1).at(0)))
        break_start1 = break_start1.add(days=1)
        break_end1 = break_end1.add(days=1)
        break_start2 = break_start2.add(days=1)
        break_end2 = break_end2.add(days=1)
    return breaks


def load_members(ws) -> tuple[str, str, MemberSet, dict[DateTime, bool]]:
    labels = [cell.value for cell in ws[1]]
    label_to_col = {label: i for i, label in enumerate(labels, start=1) if label}

    errors = []
    baseline = ws.cell(row=2, column=label_to_col["ベースライン"]).value
    team = ws.cell(row=2, column=label_to_col["チーム名"]).value

    members = MemberSet()
    rows = list(ws.iter_rows(min_row=2, max_row=3))
    names = rows[0]
    roles = rows[1]
    for i in range(1, 10):
        col = label_to_col.get(f"担当者{i}")
        if col is not None:
            name = names[col - 1].value
            if name is not None:
                role = roles[col - 1].value or "プログラマ"
                members.add(Member(name=name, role=role))

    on_off_map = {}
    i, j = find_double_datetime(ws)
    if j is None:
        errors.append("O列以降にカレンダーが見つかりませんでした。")
    else:
        rows = list(ws.iter_rows())
        i -= 1
        j -= 1
        while j < len(rows[0]):
            c = to_datetime(rows[i][j].value)
            if not isinstance(c, DateTime):
                break
            x = rows[i + 1][j].value
            on_off_map[c.at(0)] = x in ["開", "月", "火", "水", "木", "金"] or (
                x not in ["祝", "休", "土", "日"]
                and c.day_of_week not in [SATURDAY, SUNDAY]
            )
            print(
                f"{c.format('YYYY-MM-DD dddd'):<20}: {'開講' if on_off_map[c.at(0)] else '休講'}"
            )
            j += 1

    if baseline is None:
        errors.append("ベースラインの定義が見つかりません。しくしく...")
    if team is None:
        errors.append("チーム名の定義が見つかりません。しくしく...")
    if 1 > len(members):
        errors.append("担当者の定義が見当たりません。しくしく...")
    if len(errors) > 0:
        raise ValueError(errors)

    return baseline, team, members, on_off_map


def load_tasks(
    ws,
    members: MemberSet,
    now: DateTime,
    breaks: TimeRangeSet,
    on_off_map: dict[DateTime, bool],
) -> TaskSet:
    labels = None
    for row in ws.iter_rows(min_row=5, values_only=True):
        if "タスク" in row:
            labels = list(row)
            break
    if labels is None:
        raise ValueError("'タスク'というセルが見つかりません。しくしく...")
    label_to_col = {label: i for i, label in enumerate(labels) if label}
    # print(f"{label_to_col}")

    if "進捗\n(%)" in label_to_col and "進捗(%)" not in label_to_col:
        label_to_col["進捗(%)"] = label_to_col["進捗\n(%)"]
    for label in [
        "タスク",
        "進捗(%)",
        "予定開始日時",
        "予定完了日時",
        "実績開始日時",
        "実績完了日時",
        "担当者1",
    ]:
        if label not in label_to_col:
            raise ValueError(f"{label}列が見つかりません。しくしく...")
        # else:
        #    print(f"{label}: 列{label_to_col[label]}")

    print(TERM_RED, end="")
    taskset = TaskSet()
    is_warned = False
    for i, row in enumerate(
        ws.iter_rows(min_row=6, values_only=True),
        start=6,
    ):
        task_name = row[label_to_col["タスク"]]
        if task_name is None:
            continue
        task_name += f"-line{i}"
        plan_start = to_datetime(row[label_to_col["予定開始日時"]])
        if plan_start is None:
            continue
        plan_end = to_datetime(row[label_to_col["予定完了日時"]])
        if plan_end is None:
            is_warned = True
            print(f"{task_name}: 予定完了日時がありません。")
            continue
        try:
            progress = int(row[label_to_col["進捗(%)"]])
        except (TypeError, ValueError):
            progress = None
        actual_start = to_datetime(row[label_to_col["実績開始日時"]])
        actual_end = to_datetime(row[label_to_col["実績完了日時"]])
        task_members = MemberSet()
        for j in range(1, 10):
            col = label_to_col.get(f"担当者{j}")
            if col is not None:
                name = row[col]
                if name is not None:
                    m = members.find_by_name(name)
                    if m is not None:
                        task_members.add(m)
                    else:
                        print(
                            f"{task_name}: 担当者{j}の{name}さんは定義されていません。"
                        )
        if len(task_members) < 1:
            is_warned = True
            print(f"{task_name}: 恐ろしいことに誰も担当していません。")
        task = Task(
            task_name,
            progress,
            plan_start,
            plan_end,
            actual_start,
            actual_end,
            now,
            breaks,
            on_off_map,
        )
        if task.was_warned:
            is_warned = True
        taskset.add(task)
        for m in task_members:
            m.add_task(task)
            if m.was_warned:
                is_warned = True
    print(TERM_NORM, end="")
    if is_warned:
        print("\n確認したらenterを押してください。")
        input()
    return taskset


def print_progress_details(
    planned_total_seconds,
    planned_done_seconds,
    actual_total_seconds,
    actual_done_seconds,
    note="",
) -> float:
    planned_progress = 100 * planned_done_seconds / planned_total_seconds
    actual_progress = 100 * actual_done_seconds / actual_total_seconds
    actual_per_planned = (
        actual_progress / planned_progress if planned_progress > 0 else None
    )
    print(
        f"予定進捗率{note}: {planned_progress:.2f}% "
        f"({planned_done_seconds / 3600:.2f}hr/{planned_total_seconds / 3600:.2f}hr)"
    )
    print(
        f"実績進捗率{note}: {actual_progress:.2f}% "
        f"({actual_done_seconds / 3600:.2f}hr/{actual_total_seconds / 3600:.2f}hr)"
    )
    print(
        f"実績/予定 {note}: "
        + (
            f"{100 * actual_per_planned:.2f}%"
            if actual_per_planned is not None
            else "N/A"
        )
        + f" ({actual_progress:.2f}%/{planned_progress:.2f}%)"
    )
    return actual_per_planned


def gantt(trs: TaskSet, nowtt: DateTime, title: str):
    colors = {
        "plan": "rgb(100,149,237)",  # 計画  : コーンフラワーブルー
        "done": "rgb(0,255,100)",  # 完了  : 緑
        "in progress": "rgb(255,165,0)",  # 進行中: オレンジ
        "unstarted": "rgb(220,0,0)",  # 未着手: 赤
    }
    start = None
    end = None
    df = []
    for t in trs:
        if not start or t.plan_start < start:
            start = t.plan_start
        if not end or t.plan_end > end:
            end = t.plan_end
        df.append(
            dict(
                Task=t.name + "(予定)",
                Start=t.plan_start.isoformat(),
                Finish=t.plan_end.isoformat(),
                Resource="plan" if t.actual_start else "unstarted",
            )
        )
        if t.actual_end:
            if not start or t.actual_start < start:
                start = t.actual_start
            if not end or t.actual_end > end:
                end = t.actual_end
            df.append(
                dict(
                    Task=t.name,
                    Start=t.actual_start.isoformat(),
                    Finish=t.actual_end.isoformat(),
                    Resource="done",
                )
            )
        elif t.actual_start:
            if not start or t.actual_start < start:
                start = t.actual_start
            if not end or nowtt > end:
                end = nowtt
            df.append(
                dict(
                    Task=t.name,
                    Start=t.actual_start.isoformat(),
                    Finish=nowtt.isoformat(),
                    Resource="in progress",
                )
            )
    fig = ff.create_gantt(
        df,
        index_col="Resource",
        colors=colors,
        show_colorbar=True,
        group_tasks=True,
    )
    fig.update_layout(
        title=f"{title}",
        height=40 * len(df),
        width=(end - start).in_days() * 150,
        plot_bgcolor="black",
        paper_bgcolor="black",
        font=dict(color="white"),
        annotations=[
            dict(
                text=f"{title}",
                xref="paper",
                yref="paper",
                x=0.01,
                y=0.98,
                showarrow=False,
                font=dict(size=18, color="rgba(200,200,200,0.2)"),
                xanchor="left",
                yanchor="top",
            ),
            dict(
                text=f"{title}",
                xref="paper",
                yref="paper",
                x=0.5,
                y=0.5,
                showarrow=False,
                font=dict(size=18, color="rgba(200,200,200,0.2)"),
                xanchor="center",
                yanchor="middle",
            ),
            dict(
                text=f"{title}",
                xref="paper",
                yref="paper",
                x=0.98,
                y=0.02,
                showarrow=False,
                font=dict(size=18, color="rgba(200,200,200,0.2)"),
                xanchor="right",
                yanchor="bottom",
            ),
        ],
    )
    fig.update_xaxes(
        showgrid=True,
        gridwidth=1,
        dtick="D1",
        side="top" if nowtt - start < end - nowtt else "bottom",
        tickformat="%Y-%m-%d",
        gridcolor="gray",
        tickfont=dict(color="white"),
        title_font=dict(color="white"),
    )
    fig.update_yaxes(
        showgrid=True,
        gridwidth=1,
        gridcolor="gray",
        tickfont=dict(color="white"),
        title_font=dict(color="white"),
    )
    return fig


def main(xlsx: str = None, nw: str = None):
    IN_GOOGLE_COLAB = "google.colab" in sys.modules
    if not IN_GOOGLE_COLAB:
        if not xlsx and len(sys.argv) > 1:
            xlsx = sys.argv[1]
        if not nw and len(sys.argv) > 2:
            nw = sys.argv[2]
    if not xlsx:
        xlsx = "進捗管理表.xlsx"

    print(f"xlsx:{xlsx}, nw:{nw}{', Google Colab' if IN_GOOGLE_COLAB else ''}")
    nowt = penparse(nw, tz=tz_default) if nw else now(tz_default)
    wb = load_workbook(xlsx, read_only=True, data_only=True)

    ws = wb.active
    baseline, team, members, on_off_map = load_members(ws)
    print("-" * 50)
    print(f"基準: {baseline}")
    print(f"チーム名: {team}")
    for m in members:
        print("-" * 50)
        print(f"{m.name}さん")
        print(f"役割: {m.role}")
    print("-" * 50)

    breaks = make_breaks(on_off_map)
    load_tasks(ws, members, nowt, breaks, on_off_map)
    # for task in tasks:
    #     print("-" * 50)
    #     print(f"タスク: {task.name}")
    #     print(f"進捗: {task.progress}%")
    #     print(f"予定: {task.plan_start} - {task.plan_end}")
    #     print(f"実績: {task.actual_start} - {task.actual_end}")

    team_tasks = TaskSet()
    team_planned_total_seconds = 0
    team_planned_done_seconds = 0
    team_actual_total_seconds = 0
    team_actual_done_seconds = 0
    for m in members:
        team_tasks.add_tasks(m.tasks)
        (
            planned_total_seconds,
            planned_done_seconds,
            actual_total_seconds,
            actual_done_seconds,
        ) = m.tasks.total_durations()
        team_planned_total_seconds += planned_total_seconds
        team_planned_done_seconds += planned_done_seconds
        team_actual_total_seconds += actual_total_seconds
        team_actual_done_seconds += actual_done_seconds

    is_team_shown = False
    for m in members:
        base_start, base_end = m.tasks.calc_base(nowt)
        nowtt = max(base_start, nowt)
        now_days = 0
        period_days = 0
        dt = m.tasks.period_start.at(0)
        while dt < m.tasks.period_end.add(days=1).at(0):
            if on_off_map[dt]:
                period_days += 1
                if dt <= base_start:
                    now_days += 1
            dt = dt.add(days=1)
        print("=" * 80)
        print(f"{m.name}さん CS(1)")
        print("-" * 50)
        print(f"チーム名  : {team}")
        assigned_tasks = m.tasks.filter(lambda t: t.is_responsible(base_start))
        print(
            f"担当タスク: {', '.join(assigned_tasks.names()) if assigned_tasks else 'ありません。まじか、しくしく...'}"
        )
        print(f"役割      : {m.role}")
        print("-" * 50)
        print(f"{m.name}さん CS(2)")
        print("-" * 50)
        print(f"ベースライン: {baseline}")
        print(
            f"行程      : {100 * now_days / period_days:.2f}% ({now_days}日/{period_days}日 "
            + f"{nowt.format('YYYY-MM-DD HH:mmZ')}/"
            + f"{m.tasks.period_start.format('YYYY-MM-DD HH:mmZ')}/"
            + f"{m.tasks.period_end.format('YYYY-MM-DD HH:mmZ')})"
        )

        if m.role == "リーダー":
            is_team_shown = True
            print("-" * 50)
            print_progress_details(
                team_planned_total_seconds,
                team_planned_done_seconds,
                team_actual_total_seconds,
                team_actual_done_seconds,
                "(チーム)",
            )
            print("-" * 50)

        actual_per_planned = print_progress_details(*m.tasks.total_durations())

        print("コメント  : " + TERM_RED, end="")
        unstarted_tasks = m.tasks.filter(lambda t: t.is_unstarted(nowtt))
        unfinished_tasks = m.tasks.filter(lambda t: t.is_unfinished(nowtt))
        overrun_tasks = m.tasks.filter(lambda t: t.is_overrun(nowtt))
        if unstarted_tasks or unfinished_tasks or overrun_tasks:
            print("")
        if unstarted_tasks:
            print("☆ タスクが開始されていません:")
            print(
                "   "
                + "\n   ".join(
                    wljustify(f"{t.name}", unstarted_tasks.max_len_of_names)
                    + f" 予定開始日時: {t.plan_start.format('YYYY-MM-DD HH:mmZ')}"
                    + f" <= {nowtt.format('YYYY-MM-DD HH:mmZ')}"
                    for t in unstarted_tasks
                )
            )
        if unfinished_tasks:
            print("☆ タスクが完了していません:")
            print(
                "   "
                + "\n   ".join(
                    wljustify(f"{t.name}", unfinished_tasks.max_len_of_names)
                    + f" 予定終了日時: {t.plan_end.format('YYYY-MM-DD HH:mmZ')}"
                    + f" <= {nowtt.format('YYYY-MM-DD HH:mmZ')}"
                    for t in unfinished_tasks
                )
            )
        if overrun_tasks:
            print("☆ 工数が超過しています:")
            print(
                "   "
                + "\n   ".join(
                    wljustify(f"{t.name}", overrun_tasks.max_len_of_names)
                    + f" 実績工数: {(nowt - t.actual_start).in_minutes() / 60:.2f}hr"
                    + f" ({nowt.format('YYYY-MM-DD HH:mmZ')} / {t.actual_start.format('YYYY-MM-DD HH:mmZ')})"
                    + f" 予定工数: {(t.plan_end - t.plan_start).in_minutes() / 60:.2f}hr"
                    for t in overrun_tasks
                )
            )
        print(
            TERM_NORM
            + (
                "がんばります。"
                if not actual_per_planned or actual_per_planned < 50
                else "だんだん調子が出てきました。"
                if actual_per_planned < 90
                else "順調です。"
                if actual_per_planned < 120
                else "順調すぎて怖いです。"
            )
        )
        if not IN_GOOGLE_COLAB:
            print("\n確認したらenterを押してください。")
            input()

    if not is_team_shown:
        print("=" * 80)
        print_progress_details(
            team_planned_total_seconds,
            team_planned_done_seconds,
            team_actual_total_seconds,
            team_actual_done_seconds,
            "(チーム)",
        )

    if team_tasks:
        fig = gantt(
            sorted(team_tasks, key=lambda t: t.plan_start),
            nowtt,
            f"{baseline}: {xlsx} - gantt chart - Copyright (c) 2025 Fumiyuki Shimizu",
        )
        if IN_GOOGLE_COLAB:
            return fig
        fig.show()


if __name__ == "__main__":
    main()

# end of file
